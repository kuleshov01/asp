import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
import numpy as np
from datetime import datetime

from openpyxl.utils.dataframe import dataframe_to_rows

# Указываем путь к файлу и название листа
file_path = 'data/СПИСОК ДЕТЕЙ МИНСОЦ 2025.xlsx'
sheet_name = 'Список детей 2025'
year = 2025


def df_find(df):

    # Извлекаем 4-й столбец
    column_4 = df.iloc[:, 3]

    # Разделяем значения на номер и год (только строки с "/")
    # Преобразуем в строки и фильтруем по наличию '/'
    valid_values = column_4.astype(str).str.contains('/', na=False)
    split_values = column_4[valid_values].str.split('/', expand=True)
    split_values.columns = ['номер', 'год']

    # Преобразуем в числа только строки, которые можно преобразовать
    split_values['номер'] = pd.to_numeric(split_values['номер'], errors='coerce')
    split_values['год'] = pd.to_numeric(split_values['год'], errors='coerce')

    # Удаляем строки, где не удалось преобразовать (NaN)
    split_values = split_values.dropna()

    # Находим последний год и максимальный номер
    latest_year = split_values['год'].max()
    if latest_year == year:
        latest_year_data = split_values[split_values['год'] == latest_year]
        latest_number = latest_year_data['номер'].max()

        new_value = f"{latest_number}/{latest_year}"
    else:
        new_value = '1/2025'
    
    return new_value

def df_load():

    # Чтение данных из вкладки "Список детей 2025"
    # Чтение файла Excel без заголовков
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)

    orig_df = df.copy()
    
    # Преобразование всех датовых столбцов (укажите ваши колонки)
    date_columns = ['взяли на обслуживание ', 'дата ипр']  # замените на реальные названия
    for col in date_columns:
        if col in df.columns:
            # Используем errors='coerce' для преобразования некорректных значений в NaT
            df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce')

    # Чтение форматирования с помощью openpyxl
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Преобразуем заголовки в нижний регистр
    df.columns = df.columns.str.lower()

    return df, orig_df, sheet

def df_filter(df, month):
    # Фильтрация данных
    filtered_df = df[(df[month] == '+')]

    # Преобразование DataFrame в массив словарей
    records = filtered_df.to_dict('records')
    
    return records

def df_replace(df, record):
    if not record:
        return df  # Если словарь пуст
    
    # Получаем ключ и значение для поиска строки (первый элемент словаря)
    first_key, first_value = next(iter(record.items()))
    
    mask = df.iloc[:, 0] == first_value

    #print(df.columns.tolist())
    # Обновляем данные в найденной строке
    for index, (key, value) in enumerate(record.items()):

        if pd.isna(value):  # Проверка на NaN/None/NaT
            df.loc[mask, df.columns[index]] = np.nan
        #elif isinstance(value, (datetime, pd.Timestamp)):
            #breakpoint()
            # Чертовщина, в переменых данные корретны, а при сохранении timestamp меняет месяц и день
            #formatted_date = value.strftime("%d.%m.%Y")
            #df.loc[mask, df.columns[index]] = formatted_date
        else:
            df.loc[mask, df.columns[index]] = value
    
    return df


def df_save(df, sheet):
    # Сохранение данных с помощью pandas
    workbook = load_workbook(file_path)
    
    # Если страница уже существует, удаляем её
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    
    # Создаем новую страницу
    new_sheet = workbook.create_sheet(sheet_name)
       
    # Записываем заголовки (пропускаем 'Unnamed: 0')
    col_offset = 1  # Начинаем с колонки 1
    for col_num, column_name in enumerate(df.columns, col_offset):
        if "Unnamed" in str(column_name):  # Пропускаем Unnamed столбцы
            continue
        new_sheet.cell(row=2, column=col_num, value=column_name)

    # Записываем данные из DataFrame
    for row_num, row_data in enumerate(df.itertuples(index=False), 3):
        for col_num, value in enumerate(row_data, 1):
            new_sheet.cell(row=row_num, column=col_num, value=value)

    if  workbook:   
    
        for row in sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                
                # Копирование шрифта
                if cell.font:
                    new_cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color
                    )
                
                # Копирование границ
                if cell.border:
                    new_cell.border = Border(
                        left=Side(border_style=cell.border.left.border_style, color=cell.border.left.color),
                        right=Side(border_style=cell.border.right.border_style, color=cell.border.right.color),
                        top=Side(border_style=cell.border.top.border_style, color=cell.border.top.color),
                        bottom=Side(border_style=cell.border.bottom.border_style, color=cell.border.bottom.color)
                    )
                
                # Копирование заливки
                if cell.fill:
                    new_cell.fill = PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color
                    )
                
                # Копирование выравнивания
                if cell.alignment:
                    new_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent
                    )
                
                # Копирование числового формата
                new_cell.number_format = cell.number_format

        # Копируем ширину столбцов
        for col in sheet.columns:
            col_letter = col[0].column_letter
            new_sheet.column_dimensions[col_letter].width = sheet.column_dimensions[col_letter].width
        
        # Сохраняем файл
        workbook.save(file_path)

def start_exel():
    df, orig_df, sheet = df_load()
    filter_df = df_filter(df)
    new_df = df_replace(orig_df, filter_df)
    df_save(new_df, sheet)