import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def df_load():
    """
    Загрузка данных из Excel файла
    """
    file_path = "data/СПИСОК ДЕТЕЙ МИНСОЦ 2025.xlsx"
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл {file_path} не найден")
    
    # Загружаем Excel файл
    xl_file = pd.ExcelFile(file_path)
    
    # Получаем первый лист
    sheet_name = xl_file.sheet_names[0] if xl_file.sheet_names else 'Sheet1'
    
    # Читаем данные с листа
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # Загружаем книгу Excel для сохранения
    book = load_workbook(file_path)
    
    return df, df.copy(), book

def df_filter(df, month):
    """
    Фильтрация данных по месяцу
    """
    # Простая фильтрация - возвращаем все записи, где месяц не заполнен
    filtered = df[df[month].isna() | (df[month] != '!')].to_dict('records')
    return filtered

def df_replace(df, record):
    """
    Замена записи в датафрейме
    """
    # Находим индекс строки по ФИО
    index = df[df['фио '] == record['фио ']].index
    
    if len(index) > 0:
        idx = index[0]
        # Обновляем значения в строке
        for key, value in record.items():
            if key in df.columns:
                df.at[idx, key] = value
    
    return df

def df_save(df, book):
    """
    Сохранение датафрейма обратно в Excel
    """
    file_path = "data/СПИСОК ДЕТЕЙ МИНСОЦ 2025.xlsx"
    
    # Очищаем текущий лист
    sheet_name = book.sheetnames[0]
    sheet = book[sheet_name]
    sheet.delete_rows(2, sheet.max_row)  # Удаляем все строки кроме заголовка
    
    # Добавляем обновленные данные
    for r in dataframe_to_rows(df, index=False, header=False):
        sheet.append(r)
    
    # Сохраняем файл
    book.save(file_path)

def df_find(df):
    """
    Поиск значения в датафрейме (для получения префикса номера договора)
    """
    # Возвращаем тестовое значение или реализуем логику поиска
    # В оригинальном коде это использовалось для получения последнего номера договора
    return "1/2025"  # тестовое значение